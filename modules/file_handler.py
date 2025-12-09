"""
modules/file_handler.py
Módulo de carga y validación de archivos Excel
Maneja la lectura de reportes y extracción de metadata
"""

import pandas as pd
import os
import re
import streamlit as st
from datetime import datetime
from typing import Optional, Dict, Tuple, List

from config import config, messages


# ============================================================================
# EXCEPCIONES PERSONALIZADAS
# ============================================================================

class FileValidationError(Exception):
    """Excepción para errores de validación de archivos"""
    pass


class FileLoadError(Exception):
    """Excepción para errores de carga de archivos"""
    pass


# ============================================================================
# CLASE PRINCIPAL
# ============================================================================

class ExcelFileHandler:
    """Manejador de archivos Excel del reporte meteorológico"""
    
    @staticmethod
    def extraer_fechas_nombre_archivo(nombre_archivo: str) -> Tuple[Optional[str], Optional[str], Optional[int]]:
        """
        Extrae fechas del nombre de archivo siguiendo el patrón:
        reporte_disponibilidad_SGR_DDMM_DDMM.xlsx
        
        Args:
            nombre_archivo: Nombre del archivo (ej: reporte_disponibilidad_SGR_0810_1910.xlsx)
            
        Returns:
            Tupla (fecha_inicio, fecha_fin, año) en formato DD/MM/YYYY
            Retorna (None, None, None) si no encuentra el patrón
            
        Examples:
            >>> extraer_fechas_nombre_archivo("reporte_disponibilidad_SGR_0810_1910.xlsx")
            ("08/10/2025", "19/10/2025", 2025)
        """
        try:
            # Patrón: 4 dígitos_4 dígitos.xlsx
            patron = r'(\d{4})_(\d{4})\.xlsx?'
            match = re.search(patron, nombre_archivo)
            
            if not match:
                return None, None, None
            
            fecha_inicio_str = match.group(1)
            fecha_fin_str = match.group(2)
            
            # Extraer día y mes (DDMM)
            dia_inicio, mes_inicio = fecha_inicio_str[:2], fecha_inicio_str[2:]
            dia_fin, mes_fin = fecha_fin_str[:2], fecha_fin_str[2:]
            
            # Inferir año basado en mes actual
            año_actual = datetime.now().year
            mes_actual = datetime.now().month
            
            # Si el mes final es muy posterior al actual, probablemente sea del año pasado
            año = año_actual - 1 if int(mes_fin) > mes_actual + 1 else año_actual
            
            fecha_inicio = f"{dia_inicio}/{mes_inicio}/{año}"
            fecha_fin = f"{dia_fin}/{mes_fin}/{año}"
            
            return fecha_inicio, fecha_fin, año
            
        except Exception as e:
            # Si hay cualquier error, retornar None
            return None, None, None
    
    @staticmethod
    def normalizar_nombres_columnas(df: pd.DataFrame) -> pd.DataFrame:
        """
        Normaliza nombres de columnas para ser case-insensitive

        Args:
            df: DataFrame con columnas a normalizar

        Returns:
            DataFrame con columnas normalizadas
        """
        # Mapeo de nombres comunes (minúsculas -> forma correcta)
        mapeo_columnas = {
            'comentario': 'Comentario',
            'estacion': 'Estacion',
            'sensor': 'Sensor',
            'frecuencia': 'Frecuencia',
            'dz': 'DZ'
        }

        # Renombrar columnas encontradas
        columnas_nuevas = {}
        for col in df.columns:
            col_lower = col.lower()
            if col_lower in mapeo_columnas:
                columnas_nuevas[col] = mapeo_columnas[col_lower]

        if columnas_nuevas:
            df = df.rename(columns=columnas_nuevas)

        return df

    @staticmethod
    def validar_columnas(df: pd.DataFrame, columnas_requeridas: List[str], nombre_hoja: str) -> None:
        """
        Valida que el DataFrame contenga todas las columnas requeridas (case-insensitive)

        Args:
            df: DataFrame a validar
            columnas_requeridas: Lista de nombres de columnas que deben existir
            nombre_hoja: Nombre de la hoja (para mensaje de error)

        Raises:
            FileValidationError: Si faltan columnas requeridas
        """
        # Comparación case-insensitive
        columnas_df_lower = {col.lower(): col for col in df.columns}
        columnas_req_lower = {col.lower(): col for col in columnas_requeridas}

        columnas_faltantes = set(columnas_req_lower.keys()) - set(columnas_df_lower.keys())

        if columnas_faltantes:
            # Mostrar nombres originales requeridos
            nombres_faltantes = [columnas_req_lower[col] for col in columnas_faltantes]
            error_msg = (
                f"Hoja '{nombre_hoja}' - Columnas faltantes: {', '.join(nombres_faltantes)}\n"
                f"Columnas encontradas: {', '.join(df.columns)}"
            )
            raise FileValidationError(error_msg)
    
    @staticmethod
    def listar_archivos_excel(ruta_carpeta: str) -> List[str]:
        """
        Lista archivos Excel en una carpeta
        
        Args:
            ruta_carpeta: Ruta de la carpeta a explorar
            
        Returns:
            Lista de nombres de archivo ordenados por fecha (más reciente primero)
        """
        if not os.path.exists(ruta_carpeta):
            return []
        
        try:
            archivos = [
                f for f in os.listdir(ruta_carpeta)
                if f.endswith(('.xlsx', '.xls')) and not f.startswith('~$')  # Excluir temporales
            ]
            
            # Ordenar por fecha de modificación (más reciente primero)
            archivos_con_fecha = []
            for archivo in archivos:
                ruta_completa = os.path.join(ruta_carpeta, archivo)
                fecha_mod = os.path.getmtime(ruta_completa)
                archivos_con_fecha.append((archivo, fecha_mod))
            
            # Ordenar por fecha descendente
            archivos_ordenados = [
                archivo for archivo, _ in sorted(archivos_con_fecha, key=lambda x: x[1], reverse=True)
            ]
            
            return archivos_ordenados
            
        except Exception as e:
            return []
    
    @classmethod
    @st.cache_data(ttl=3600)  # Cache por 1 hora
    def cargar_excel(_cls, archivo_path) -> Optional[Dict]:
        """
        Carga archivo Excel con validación completa
        
        Args:
            archivo_path: Ruta al archivo Excel o objeto UploadedFile de Streamlit
            
        Returns:
            Dict con estructura:
            {
                'estaciones': DataFrame,
                'sensores': DataFrame,
                'variables': DataFrame,
                'metadata': {
                    'nombre_archivo': str,
                    'fecha_inicio': str,
                    'fecha_fin': str,
                    'año': int,
                    'fecha_carga': str
                }
            }
            
        Raises:
            FileLoadError: Si hay error al cargar el archivo
            FileValidationError: Si falta alguna columna requerida
        """
        try:
            # Cargar las 3 hojas requeridas
            df_estaciones = pd.read_excel(archivo_path, sheet_name=config.SHEET_ESTACIONES)
            df_sensores = pd.read_excel(archivo_path, sheet_name=config.SHEET_SENSORES)
            df_variables = pd.read_excel(archivo_path, sheet_name=config.SHEET_VARIABLES)

            # Normalizar nombres de columnas (comentario -> Comentario, etc.)
            df_estaciones = _cls.normalizar_nombres_columnas(df_estaciones)
            df_sensores = _cls.normalizar_nombres_columnas(df_sensores)
            df_variables = _cls.normalizar_nombres_columnas(df_variables)

            # Validar columnas de cada hoja (ahora case-insensitive)
            _cls.validar_columnas(
                df_estaciones,
                config.REQUIRED_COLUMNS[config.SHEET_ESTACIONES],
                config.SHEET_ESTACIONES
            )
            _cls.validar_columnas(
                df_sensores,
                config.REQUIRED_COLUMNS[config.SHEET_SENSORES],
                config.SHEET_SENSORES
            )
            _cls.validar_columnas(
                df_variables,
                config.REQUIRED_COLUMNS[config.SHEET_VARIABLES],
                config.SHEET_VARIABLES
            )
            
            # Extraer nombre del archivo
            if isinstance(archivo_path, str):
                nombre_archivo = os.path.basename(archivo_path)
            else:
                # Es un UploadedFile de Streamlit
                nombre_archivo = archivo_path.name
            
            # Extraer fechas del nombre
            fecha_inicio, fecha_fin, año = _cls.extraer_fechas_nombre_archivo(nombre_archivo)
            
            # Construir resultado
            resultado = {
                'estaciones': df_estaciones,
                'sensores': df_sensores,
                'variables': df_variables,
                'metadata': {
                    'nombre_archivo': nombre_archivo,
                    'fecha_inicio': fecha_inicio,
                    'fecha_fin': fecha_fin,
                    'año': año,
                    'fecha_carga': datetime.now().strftime(config.DATETIME_FORMAT),
                    'num_estaciones': len(df_estaciones),
                    'num_sensores': len(df_sensores),
                    'num_variables': len(df_variables)
                }
            }
            
            return resultado
            
        except FileValidationError:
            # Re-lanzar errores de validación
            raise
            
        except Exception as e:
            # Cualquier otro error se convierte en FileLoadError
            raise FileLoadError(f"Error al cargar el archivo: {str(e)}")
    
    @staticmethod
    def exportar_csv(df: pd.DataFrame, prefijo: str = "export") -> bytes:
        """
        Exporta DataFrame a CSV con encoding UTF-8
        
        Args:
            df: DataFrame a exportar
            prefijo: Prefijo para el nombre del archivo (no se usa aquí, solo para referencia)
            
        Returns:
            Bytes del CSV codificado en UTF-8 con BOM (para Excel)
        """
        # Usar utf-8-sig para que Excel abra correctamente con tildes
        csv_data = df.to_csv(index=False).encode('utf-8-sig')
        return csv_data
    
    @staticmethod
    def crear_nombre_descarga(prefijo: str, extension: str = "csv") -> str:
        """
        Crea nombre de archivo para descarga con timestamp
        
        Args:
            prefijo: Prefijo del archivo (ej: 'estaciones', 'sensores')
            extension: Extensión del archivo (default: 'csv')
            
        Returns:
            Nombre de archivo con formato: prefijo_YYYYMMDD_HHMM.extension
            
        Examples:
            >>> crear_nombre_descarga('estaciones')
            'estaciones_20251104_1530.csv'
        """
        timestamp = datetime.now().strftime("%Y%m%d_%H%M")
        return f"{prefijo}_{timestamp}.{extension}"

    @staticmethod
    def exportar_metricas_excel(
        metricas_globales: Dict,
        df_estaciones: pd.DataFrame,
        df_sensores: pd.DataFrame,
        df_variables: pd.DataFrame,
        dz_seleccionada: str = "Todas"
    ) -> bytes:
        """
        Exporta métricas a Excel con múltiples hojas

        Args:
            metricas_globales: Diccionario con métricas globales
            df_estaciones: DataFrame de estaciones procesado
            df_sensores: DataFrame de sensores procesado
            df_variables: DataFrame de variables procesado
            dz_seleccionada: DZ filtrada o "Todas"

        Returns:
            Bytes del archivo Excel
        """
        from io import BytesIO
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment

        # Crear workbook
        wb = Workbook()
        wb.remove(wb.active)  # Eliminar hoja por defecto

        # Hoja 1: Métricas Globales
        ws_global = wb.create_sheet("Métricas Globales")
        ws_global.append(["Dashboard Meteorológico SGR - Métricas Globales"])
        ws_global.append(["Fecha de generación:", datetime.now().strftime("%d/%m/%Y %H:%M")])
        ws_global.append(["Filtro DZ:", dz_seleccionada])
        ws_global.append([])  # Línea en blanco

        # Encabezados
        ws_global.append(["Métrica", "Valor"])
        ws_global.append(["Disponibilidad Promedio", f"{metricas_globales.get('disponibilidad_promedio', 0):.1f}%"])
        ws_global.append(["Total Estaciones", metricas_globales.get('total_estaciones', 0)])
        ws_global.append(["Estaciones Críticas (<80%)", metricas_globales.get('estaciones_criticas', 0)])
        ws_global.append(["% Red Crítico", f"{(metricas_globales.get('estaciones_criticas', 0) / metricas_globales.get('total_estaciones', 1) * 100):.1f}%" if metricas_globales.get('total_estaciones', 0) > 0 else "0%"])
        ws_global.append(["Anomalías (>100%)", metricas_globales.get('estaciones_anomalias', 0)])
        ws_global.append(["DZ Afectadas", metricas_globales.get('dz_afectadas', 0)])

        # Hoja 2: Métricas Estación
        ws_estacion = wb.create_sheet("Métricas Estación")
        ws_estacion.append(["Métricas por Estación"])
        ws_estacion.append(["Filtro DZ:", dz_seleccionada])
        ws_estacion.append([])

        total_est = len(df_estaciones)
        operativas_est = (df_estaciones['disponibilidad'] >= 80).sum()
        inoperativas_est = (df_estaciones['disponibilidad'] == 0).sum()
        parcialmente_est = total_est - operativas_est - inoperativas_est

        # Estaciones con incidencias activas
        if 'estado_inci' in df_estaciones.columns:
            estados_activos = ['nueva', 'recurrente']
            con_incidencias = df_estaciones[
                df_estaciones['estado_inci'].str.lower().isin(estados_activos)
            ].shape[0]
        else:
            con_incidencias = 0

        disponibilidad_prom_est = df_estaciones['disponibilidad'].mean()

        ws_estacion.append(["Métrica", "Cantidad", "Porcentaje"])
        ws_estacion.append(["Total Estaciones", total_est, "100.0%"])
        ws_estacion.append(["Estaciones Operativas (≥80%)", operativas_est, f"{(operativas_est/total_est*100):.1f}%" if total_est > 0 else "0%"])
        ws_estacion.append(["Estaciones Parcialmente Operativas (>0% y <80%)", parcialmente_est, f"{(parcialmente_est/total_est*100):.1f}%" if total_est > 0 else "0%"])
        ws_estacion.append(["Estaciones Inoperativas (=0%)", inoperativas_est, f"{(inoperativas_est/total_est*100):.1f}%" if total_est > 0 else "0%"])
        ws_estacion.append(["Con Incidencias Activas", con_incidencias, f"{(con_incidencias/total_est*100):.1f}%" if total_est > 0 else "0%"])
        ws_estacion.append(["Disponibilidad Promedio", f"{disponibilidad_prom_est:.1f}%", ""])

        # Hoja 3: Métricas Sensor
        ws_sensor = wb.create_sheet("Métricas Sensor")
        ws_sensor.append(["Métricas por Sensor/Equipamiento"])
        ws_sensor.append(["Filtro DZ:", dz_seleccionada])
        ws_sensor.append([])

        total_sen = len(df_sensores)
        operativos_sen = (df_sensores['disponibilidad'] >= 80).sum()
        inoperativos_sen = (df_sensores['disponibilidad'] == 0).sum()
        parcialmente_sen = total_sen - operativos_sen - inoperativos_sen
        criticos_sen = (df_sensores['disponibilidad'] < 80).sum()

        ws_sensor.append(["Métrica", "Cantidad", "Porcentaje"])
        ws_sensor.append(["Total Sensores", total_sen, "100.0%"])
        ws_sensor.append(["Sensores Operativos (≥80%)", operativos_sen, f"{(operativos_sen/total_sen*100):.1f}%" if total_sen > 0 else "0%"])
        ws_sensor.append(["Sensores Parcialmente Operativos (>0% y <80%)", parcialmente_sen, f"{(parcialmente_sen/total_sen*100):.1f}%" if total_sen > 0 else "0%"])
        ws_sensor.append(["Sensores Inoperativos (=0%)", inoperativos_sen, f"{(inoperativos_sen/total_sen*100):.1f}%" if total_sen > 0 else "0%"])
        ws_sensor.append(["Críticos (<80%)", criticos_sen, f"{(criticos_sen/total_sen*100):.1f}%" if total_sen > 0 else "0%"])

        # Hoja 4: Métricas Variable
        ws_variable = wb.create_sheet("Métricas Variable")
        ws_variable.append(["Métricas por Variable Meteorológica"])
        ws_variable.append(["Filtro DZ:", dz_seleccionada])
        ws_variable.append([])

        total_var = len(df_variables)
        datos_esperados = df_variables['Datos_esperados'].sum()
        datos_recibidos = df_variables['datos_recibidos'].sum()
        datos_faltantes = datos_esperados - datos_recibidos
        datos_erroneos = df_variables['Datos_flag_M'].sum()

        ws_variable.append(["Métrica", "Cantidad", "Porcentaje"])
        ws_variable.append(["Total Variables Registradas", total_var, ""])
        ws_variable.append(["Datos Esperados", datos_esperados, "100.0%"])
        ws_variable.append(["Datos Recibidos", datos_recibidos, f"{(datos_recibidos/datos_esperados*100):.1f}%" if datos_esperados > 0 else "0%"])
        ws_variable.append(["Datos Faltantes", datos_faltantes, f"{(datos_faltantes/datos_esperados*100):.1f}%" if datos_esperados > 0 else "0%"])
        ws_variable.append(["Datos Erróneos (Flag M)", datos_erroneos, f"{(datos_erroneos/datos_recibidos*100):.1f}%" if datos_recibidos > 0 else "0%"])

        # Aplicar estilos a todas las hojas
        for ws in [ws_global, ws_estacion, ws_sensor, ws_variable]:
            # Título en negrita y fondo
            ws['A1'].font = Font(bold=True, size=14)
            ws['A1'].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            ws['A1'].font = Font(bold=True, size=14, color="FFFFFF")

            # Ajustar ancho de columnas
            ws.column_dimensions['A'].width = 40
            ws.column_dimensions['B'].width = 20
            if ws.max_column >= 3:
                ws.column_dimensions['C'].width = 15

        # Guardar en BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        return output.getvalue()

    @staticmethod
    def exportar_metricas_excel_consolidado(
        metricas_globales: Dict,
        df_estaciones: pd.DataFrame,
        df_sensores: pd.DataFrame,
        df_variables: pd.DataFrame,
        dz_seleccionada: str = "Todas"
    ) -> bytes:
        """
        Exporta todas las métricas en UNA SOLA hoja consolidada con columna Módulo

        Args:
            metricas_globales: Diccionario con métricas globales
            df_estaciones: DataFrame de estaciones procesado
            df_sensores: DataFrame de sensores procesado
            df_variables: DataFrame de variables procesado
            dz_seleccionada: DZ filtrada o "Todas"

        Returns:
            Bytes del archivo Excel
        """
        from io import BytesIO
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        # Crear lista consolidada de métricas
        metricas_consolidadas = []

        # Calcular todas las métricas
        total_est = len(df_estaciones)
        operativas_est = (df_estaciones['disponibilidad'] >= 80).sum()
        inoperativas_est = (df_estaciones['disponibilidad'] == 0).sum()
        parcialmente_est = total_est - operativas_est - inoperativas_est

        if 'estado_inci' in df_estaciones.columns:
            estados_activos = ['nueva', 'recurrente']
            con_incidencias = df_estaciones[
                df_estaciones['estado_inci'].str.lower().isin(estados_activos)
            ].shape[0]
        else:
            con_incidencias = 0

        disponibilidad_prom_est = df_estaciones['disponibilidad'].mean()

        total_sen = len(df_sensores)
        operativos_sen = (df_sensores['disponibilidad'] >= 80).sum()
        inoperativos_sen = (df_sensores['disponibilidad'] == 0).sum()
        parcialmente_sen = total_sen - operativos_sen - inoperativos_sen
        criticos_sen = (df_sensores['disponibilidad'] < 80).sum()

        total_var = len(df_variables)
        datos_esperados = df_variables['Datos_esperados'].sum()
        datos_recibidos = df_variables['datos_recibidos'].sum()
        datos_faltantes = datos_esperados - datos_recibidos
        datos_erroneos = df_variables['Datos_flag_M'].sum()

        # MÓDULO: Global
        metricas_consolidadas.append(["Global", "Disponibilidad Promedio", f"{metricas_globales.get('disponibilidad_promedio', 0):.1f}%", "", dz_seleccionada])
        metricas_consolidadas.append(["Global", "Total Estaciones", metricas_globales.get('total_estaciones', 0), "100.0%", dz_seleccionada])
        metricas_consolidadas.append(["Global", "Estaciones Críticas (<80%)", metricas_globales.get('estaciones_criticas', 0), f"{(metricas_globales.get('estaciones_criticas', 0) / metricas_globales.get('total_estaciones', 1) * 100):.1f}%" if metricas_globales.get('total_estaciones', 0) > 0 else "0%", dz_seleccionada])
        metricas_consolidadas.append(["Global", "Anomalías (>100%)", metricas_globales.get('estaciones_anomalias', 0), "", dz_seleccionada])
        metricas_consolidadas.append(["Global", "DZ Afectadas", metricas_globales.get('dz_afectadas', 0), "", dz_seleccionada])

        # MÓDULO: Estación
        metricas_consolidadas.append(["Estación", "Total Estaciones", total_est, "100.0%", dz_seleccionada])
        metricas_consolidadas.append(["Estación", "Estaciones Operativas (≥80%)", operativas_est, f"{(operativas_est/total_est*100):.1f}%" if total_est > 0 else "0%", dz_seleccionada])
        metricas_consolidadas.append(["Estación", "Estaciones Parcialmente Operativas (>0% y <80%)", parcialmente_est, f"{(parcialmente_est/total_est*100):.1f}%" if total_est > 0 else "0%", dz_seleccionada])
        metricas_consolidadas.append(["Estación", "Estaciones Inoperativas (=0%)", inoperativas_est, f"{(inoperativas_est/total_est*100):.1f}%" if total_est > 0 else "0%", dz_seleccionada])
        metricas_consolidadas.append(["Estación", "Con Incidencias Activas", con_incidencias, f"{(con_incidencias/total_est*100):.1f}%" if total_est > 0 else "0%", dz_seleccionada])
        metricas_consolidadas.append(["Estación", "Disponibilidad Promedio", f"{disponibilidad_prom_est:.1f}%", "", dz_seleccionada])

        # MÓDULO: Sensor
        metricas_consolidadas.append(["Sensor", "Total Sensores", total_sen, "100.0%", dz_seleccionada])
        metricas_consolidadas.append(["Sensor", "Sensores Operativos (≥80%)", operativos_sen, f"{(operativos_sen/total_sen*100):.1f}%" if total_sen > 0 else "0%", dz_seleccionada])
        metricas_consolidadas.append(["Sensor", "Sensores Parcialmente Operativos (>0% y <80%)", parcialmente_sen, f"{(parcialmente_sen/total_sen*100):.1f}%" if total_sen > 0 else "0%", dz_seleccionada])
        metricas_consolidadas.append(["Sensor", "Sensores Inoperativos (=0%)", inoperativos_sen, f"{(inoperativos_sen/total_sen*100):.1f}%" if total_sen > 0 else "0%", dz_seleccionada])
        metricas_consolidadas.append(["Sensor", "Críticos (<80%)", criticos_sen, f"{(criticos_sen/total_sen*100):.1f}%" if total_sen > 0 else "0%", dz_seleccionada])

        # MÓDULO: Variable
        metricas_consolidadas.append(["Variable", "Total Variables Registradas", total_var, "", dz_seleccionada])
        metricas_consolidadas.append(["Variable", "Datos Esperados", datos_esperados, "100.0%", dz_seleccionada])
        metricas_consolidadas.append(["Variable", "Datos Recibidos", datos_recibidos, f"{(datos_recibidos/datos_esperados*100):.1f}%" if datos_esperados > 0 else "0%", dz_seleccionada])
        metricas_consolidadas.append(["Variable", "Datos Faltantes", datos_faltantes, f"{(datos_faltantes/datos_esperados*100):.1f}%" if datos_esperados > 0 else "0%", dz_seleccionada])
        metricas_consolidadas.append(["Variable", "Datos Erróneos (Flag M)", datos_erroneos, f"{(datos_erroneos/datos_recibidos*100):.1f}%" if datos_recibidos > 0 else "0%", dz_seleccionada])

        # Crear DataFrame consolidado
        df_consolidado = pd.DataFrame(
            metricas_consolidadas,
            columns=["Módulo", "Métrica", "Valor", "Porcentaje", "DZ"]
        )

        # Exportar a Excel con formato
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df_consolidado.to_excel(writer, sheet_name='Métricas Consolidadas', index=False)

            # Obtener workbook y worksheet
            workbook = writer.book
            worksheet = writer.sheets['Métricas Consolidadas']

            # Estilos
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=11)
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            # Aplicar estilos al encabezado
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border

            # Aplicar bordes a todas las celdas
            for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=1, max_col=5):
                for cell in row:
                    cell.border = border

            # Ajustar ancho de columnas
            worksheet.column_dimensions['A'].width = 15  # Módulo
            worksheet.column_dimensions['B'].width = 50  # Métrica
            worksheet.column_dimensions['C'].width = 20  # Valor
            worksheet.column_dimensions['D'].width = 15  # Porcentaje
            worksheet.column_dimensions['E'].width = 20  # DZ

            # Agregar metadata en las primeras filas
            worksheet.insert_rows(1, 3)
            worksheet['A1'] = "Dashboard Meteorológico SGR - Métricas Consolidadas"
            worksheet['A1'].font = Font(bold=True, size=14)
            worksheet['A2'] = f"Fecha de generación: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
            worksheet['A3'] = f"Filtro DZ: {dz_seleccionada}"

        output.seek(0)
        return output.getvalue()


# ============================================================================
# FUNCIONES DE UTILIDAD
# ============================================================================

def verificar_estructura_carpeta(ruta: str) -> Dict[str, any]:
    """
    Verifica el estado de una carpeta de reportes
    
    Args:
        ruta: Ruta de la carpeta a verificar
        
    Returns:
        Dict con información:
        {
            'existe': bool,
            'num_archivos': int,
            'archivos': List[str],
            'ultimo_modificado': Optional[str]
        }
    """
    resultado = {
        'existe': False,
        'num_archivos': 0,
        'archivos': [],
        'ultimo_modificado': None
    }
    
    if not os.path.exists(ruta):
        return resultado
    
    resultado['existe'] = True
    
    handler = ExcelFileHandler()
    archivos = handler.listar_archivos_excel(ruta)
    
    resultado['num_archivos'] = len(archivos)
    resultado['archivos'] = archivos
    
    if archivos:
        ruta_ultimo = os.path.join(ruta, archivos[0])
        fecha_mod = datetime.fromtimestamp(os.path.getmtime(ruta_ultimo))
        resultado['ultimo_modificado'] = fecha_mod.strftime(config.DATETIME_FORMAT)
    
    return resultado